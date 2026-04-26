from __future__ import annotations

from pathlib import Path
from typing import Any

from schemas import (
    EXCEL_TOOL_DEPENDENCY_ERROR,
    EXCEL_TOOL_ERROR,
    EXCEL_TOOL_FILE_NOT_FOUND,
    EXCEL_TOOL_SHEET_EXISTS,
    EXCEL_TOOL_SHEET_NOT_FOUND,
    TOOL_ARGUMENT_ERROR,
    ToolResult,
    build_error,
)
from tools.tools import BaseTool, build_tool_output
from utils.env_util.runtime_env import get_task_runtime_dir


class ExcelTool(BaseTool):
    name = "excel"
    description = (
        "Inspect, read, and update Excel workbooks (.xlsx/.xlsm). "
        "Actions: inspect (list sheets with row/column dimensions), "
        "read_sheet (read up to max_rows rows; returns cached values, not formula text), "
        "write_sheet (write rows to a sheet; creates sheet if absent; fails if sheet exists and replace_sheet=false), "
        "append_rows (append rows after existing data; creates sheet/file if absent). "
        "Relative paths resolve inside the task workspace directory."
    )
    parameters = {
        "type": "object",
        "properties": {
            "action": {
                "type": "string",
                "description": (
                    "inspect: list sheets and dimensions. "
                    "read_sheet: read rows (returns cached values, not formula text). "
                    "write_sheet: write rows to a sheet (creates if absent; fails if exists and replace_sheet=false). "
                    "append_rows: append rows after existing data (creates sheet/file if absent)."
                ),
                "enum": ["inspect", "read_sheet", "write_sheet", "append_rows"],
            },
            "path": {
                "type": "string",
                "description": (
                    "Path to the Excel workbook (.xlsx/.xlsm). "
                    "Relative paths resolve inside the task workspace. "
                    "Created automatically for write_sheet and append_rows if absent."
                ),
            },
            "sheet_name": {
                "type": "string",
                "description": (
                    "Target worksheet name. Defaults to the active sheet when omitted. "
                    "Created automatically for write_sheet and append_rows if it does not exist."
                ),
            },
            "max_rows": {
                "type": "integer",
                "description": "Max rows to return for read_sheet. Defaults to 50, max 500.",
                "default": 50,
            },
            "rows": {
                "type": "array",
                "description": (
                    "Required for write_sheet and append_rows. "
                    "Array of row arrays; cell values may be strings, numbers, booleans, or null. "
                    "Non-primitive values are coerced to strings."
                ),
                "items": {
                    "type": "array",
                    "items": {},
                },
            },
            "replace_sheet": {
                "type": "boolean",
                "description": (
                    "For write_sheet only. When true (default), deletes and recreates the sheet before writing. "
                    "When false, fails if the sheet already exists."
                ),
                "default": True,
            },
        },
        "required": ["action", "path"],
        "additionalProperties": False,
    }

    def run(self, arguments: dict[str, object]) -> ToolResult:
        action = str(arguments.get("action", "")).strip().lower()
        path_value = str(arguments.get("path", "")).strip()
        sheet_name = str(arguments.get("sheet_name", "")).strip() or None

        if action not in {"inspect", "read_sheet", "write_sheet", "append_rows"}:
            error = build_error(
                TOOL_ARGUMENT_ERROR,
                "Excel tool action must be inspect, read_sheet, write_sheet, or append_rows.",
            )
            return self._error_result(error)
        if not path_value:
            error = build_error(TOOL_ARGUMENT_ERROR, "Excel tool requires a non-empty path.")
            return self._error_result(error)

        path = self._resolve_target_path(path_value)
        try:
            if action == "inspect":
                return self._inspect_workbook(path)
            if action == "read_sheet":
                max_rows = self._normalize_max_rows(arguments.get("max_rows", 50))
                return self._read_sheet(path, sheet_name=sheet_name, max_rows=max_rows)

            rows = self._normalize_rows(arguments.get("rows"))
            if isinstance(rows, ToolResult):
                return rows

            if action == "write_sheet":
                replace_sheet = bool(arguments.get("replace_sheet", True))
                return self._write_sheet(
                    path,
                    sheet_name=sheet_name,
                    rows=rows,
                    replace_sheet=replace_sheet,
                )

            return self._append_rows(path, sheet_name=sheet_name, rows=rows)
        except Exception as exc:
            error = build_error(EXCEL_TOOL_ERROR, f"Excel tool failed: {exc}")
            return self._error_result(error)

    def _inspect_workbook(self, path: Path) -> ToolResult:
        workbook = self._load_workbook(path, read_only=True)
        try:
            sheets = []
            for sheet in workbook.worksheets:
                sheets.append(
                    {
                        "name": sheet.title,
                        "max_row": sheet.max_row,
                        "max_column": sheet.max_column,
                    }
                )
            return ToolResult(
                output=build_tool_output(
                    success=True,
                    data={
                        "action": "inspect",
                        "path": str(path),
                        "sheet_count": len(sheets),
                        "active_sheet": workbook.active.title,
                        "sheets": sheets,
                    },
                ),
                success=True,
            )
        finally:
            workbook.close()

    def _read_sheet(
        self,
        path: Path,
        *,
        sheet_name: str | None,
        max_rows: int,
    ) -> ToolResult:
        workbook = self._load_workbook(path, read_only=True)
        try:
            sheet = self._get_sheet(workbook, sheet_name)
            rows: list[list[Any]] = []
            for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
                if row_index >= max_rows:
                    break
                rows.append(list(row))
            return ToolResult(
                output=build_tool_output(
                    success=True,
                    data={
                        "action": "read_sheet",
                        "path": str(path),
                        "sheet_name": sheet.title,
                        "max_row": sheet.max_row,
                        "max_column": sheet.max_column,
                        "returned_row_count": len(rows),
                        "rows": rows,
                    },
                ),
                success=True,
            )
        finally:
            workbook.close()

    def _write_sheet(
        self,
        path: Path,
        *,
        sheet_name: str | None,
        rows: list[list[Any]],
        replace_sheet: bool,
    ) -> ToolResult:
        workbook = self._load_or_create_workbook(path)
        target_sheet_name = sheet_name or workbook.active.title
        try:
            if target_sheet_name in workbook.sheetnames and replace_sheet:
                existing = workbook[target_sheet_name]
                workbook.remove(existing)
                sheet = workbook.create_sheet(title=target_sheet_name)
            elif target_sheet_name in workbook.sheetnames:
                error = build_error(
                    EXCEL_TOOL_SHEET_EXISTS,
                    f"Excel sheet already exists: {target_sheet_name}. Set replace_sheet=true to overwrite it.",
                )
                return self._error_result(error)
            elif len(workbook.sheetnames) == 1 and workbook.active.max_row == 1 and workbook.active.max_column == 1 and workbook.active["A1"].value is None:
                sheet = workbook.active
                sheet.title = target_sheet_name
            else:
                sheet = workbook.create_sheet(title=target_sheet_name)

            for row in rows:
                sheet.append(row)
            self._save_workbook(workbook, path)
            return ToolResult(
                output=build_tool_output(
                    success=True,
                    data={
                        "action": "write_sheet",
                        "path": str(path),
                        "sheet_name": sheet.title,
                        "row_count": len(rows),
                    },
                ),
                success=True,
            )
        finally:
            workbook.close()

    def _append_rows(
        self,
        path: Path,
        *,
        sheet_name: str | None,
        rows: list[list[Any]],
    ) -> ToolResult:
        workbook = self._load_or_create_workbook(path)
        target_sheet_name = sheet_name or workbook.active.title
        try:
            if target_sheet_name in workbook.sheetnames:
                sheet = workbook[target_sheet_name]
            elif len(workbook.sheetnames) == 1 and workbook.active.max_row == 1 and workbook.active.max_column == 1 and workbook.active["A1"].value is None:
                sheet = workbook.active
                sheet.title = target_sheet_name
                sheet.delete_rows(1)  # remove phantom empty row openpyxl adds to new workbooks
            else:
                sheet = workbook.create_sheet(title=target_sheet_name)
            for row in rows:
                sheet.append(row)
            self._save_workbook(workbook, path)
            return ToolResult(
                output=build_tool_output(
                    success=True,
                    data={
                        "action": "append_rows",
                        "path": str(path),
                        "sheet_name": sheet.title,
                        "rows_appended": len(rows),
                        "max_row": sheet.max_row,
                    },
                ),
                success=True,
            )
        finally:
            workbook.close()

    @staticmethod
    def _normalize_max_rows(value: object) -> int:
        try:
            max_rows = int(value)
        except (TypeError, ValueError):
            max_rows = 50
        return max(1, min(max_rows, 500))

    def _normalize_rows(self, value: object) -> list[list[Any]] | ToolResult:
        if not isinstance(value, list) or not value:
            error = build_error(
                TOOL_ARGUMENT_ERROR,
                "Excel tool requires a non-empty rows array for write_sheet and append_rows.",
            )
            return self._error_result(error)
        normalized: list[list[Any]] = []
        for row in value:
            if not isinstance(row, list):
                error = build_error(
                    TOOL_ARGUMENT_ERROR,
                    "Excel tool rows must be an array of row arrays.",
                )
                return self._error_result(error)
            normalized.append([self._normalize_cell(cell) for cell in row])
        return normalized

    @staticmethod
    def _normalize_cell(value: Any) -> Any:
        if value is None or isinstance(value, (bool, int, float, str)):
            return value
        return str(value)

    def _load_workbook(self, path: Path, *, read_only: bool) -> Any:
        if not path.exists():
            raise build_error(EXCEL_TOOL_FILE_NOT_FOUND, f"Excel workbook not found: {path}")
        openpyxl = self._require_openpyxl()
        return openpyxl.load_workbook(path, read_only=read_only, data_only=True)

    def _load_or_create_workbook(self, path: Path) -> Any:
        openpyxl = self._require_openpyxl()
        path.parent.mkdir(parents=True, exist_ok=True)
        if path.exists():
            return openpyxl.load_workbook(path)
        return openpyxl.Workbook()

    @staticmethod
    def _save_workbook(workbook: Any, path: Path) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(path)

    @staticmethod
    def _get_sheet(workbook: Any, sheet_name: str | None) -> Any:
        if sheet_name is None:
            return workbook.active
        try:
            return workbook[sheet_name]
        except KeyError as exc:
            available = ", ".join(workbook.sheetnames) or "<none>"
            raise build_error(
                EXCEL_TOOL_SHEET_NOT_FOUND,
                f"Unknown Excel sheet: {sheet_name}. Available sheets: {available}",
            ) from exc

    @staticmethod
    def _resolve_target_path(path_value: str) -> Path:
        target_path = Path(path_value).expanduser()
        if target_path.is_absolute():
            return target_path
        try:
            return get_task_runtime_dir() / target_path
        except RuntimeError:
            pass
        return target_path

    @staticmethod
    def _require_openpyxl() -> Any:
        try:
            import openpyxl
        except ModuleNotFoundError as exc:
            raise build_error(
                EXCEL_TOOL_DEPENDENCY_ERROR,
                "Excel tool requires the `openpyxl` package to be installed.",
            ) from exc
        return openpyxl

    @staticmethod
    def _error_result(error) -> ToolResult:
        return ToolResult(
            output=build_tool_output(success=False, error=error),
            success=False,
            error=error,
        )
